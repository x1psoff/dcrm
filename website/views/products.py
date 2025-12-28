"""Функции для работы с продуктами"""
from django.shortcuts import render, redirect, get_object_or_404
from django.contrib import messages
from django.contrib.auth.decorators import login_required
from django.http import HttpResponse, JsonResponse
from django.db.models import Q
from django.conf import settings
from openpyxl import load_workbook
from io import BytesIO
from decimal import Decimal, InvalidOperation
from ..models import RecordProduct, Record, Product, Category, ProductCustomField
from ..forms import ProductFilterForm
from django.views.decorators.http import require_http_methods
from django.contrib.auth.decorators import user_passes_test


def export_products(request, pk):
    record = get_object_or_404(Record, id=pk)

    # Загружаем шаблон Excel
    import os
    template_path = os.path.join(settings.MEDIA_ROOT, 'templates', 'комплектация Кухни.xlsx')
    wb = load_workbook(template_path)
    ws = wb.active

    # Сопоставление типов петель со строками в Excel
    petli_mapping = {
        'накладная': 28,
        'пружинная': 29,
        'смежная': 30,
        'вкладная': 31,
        'фальш-планка': 32,
        '155гр.': 33,
        'для алюм рамок': 34
    }

    # Сопоставление типов направляющих со строками в Excel
    naprav_mapping = {
        'частичного выдвижения': 35,
        'скрыт.мон.дов': 36,
        'скрыт.мон.типон': 37,
        'полного выдвижения': 44
    }

    # Загружаем все RecordProduct одним запросом для избежания N+1
    record_products_dict = {
        rp.product_id: rp for rp in RecordProduct.objects.filter(record=record).select_related('product')
    }

    # Обрабатываем продукты записи с оптимизацией запросов
    for product in record.products.select_related('category').all():
        record_product = record_products_dict.get(product.id)
        if not record_product:
            continue
        
        buyer = record_product.buyer
        quantity = record_product.quantity

        # Определяем категорию продукта
        category_name = product.category.name.lower() if product.category else ""

        # Обрабатываем петли
        if 'петл' in category_name and product.mounting_type in petli_mapping:
            row = petli_mapping[product.mounting_type]
            ws[f'D{row}'] = quantity
            ws[f'E{row}'] = product.name
            if buyer == 'Юра':
                ws[f'G{row}'] = product.our_price * quantity
            else:
                ws[f'H{row}'] = product.our_price * quantity

        # Обрабатываем направляющие
        elif 'направля' in category_name:
            naprav_type = None
            if product.runner_size:
                naprav_type = product.runner_size
            elif product.response_type:
                naprav_type = product.response_type

            if naprav_type in naprav_mapping:
                row = naprav_mapping[naprav_type]
                ws[f'D{row}'] = quantity
                ws[f'E{row}'] = product.name
                if buyer == 'Юра':
                    ws[f'G{row}'] = product.our_price * quantity
                else:
                    ws[f'H{row}'] = product.our_price * quantity
            else:
                row = 48
                ws[f'B{row}'] = naprav_type
                ws[f'D{row}'] = quantity
                ws[f'E{row}'] = product.name
                if buyer == 'Юра':
                    ws[f'G{row}'] = product.our_price * quantity
                else:
                    ws[f'H{row}'] = product.our_price * quantity

        # Обрабатываем другие продукты
        else:
            row = 48
            while ws[f'B{row}'].value is not None:
                row += 1
            ws[f'B{row}'] = category_name
            ws[f'D{row}'] = quantity
            ws[f'E{row}'] = product.name
            if buyer == 'Юра':
                ws[f'G{row}'] = product.our_price * quantity
            else:
                ws[f'H{row}'] = product.our_price * quantity

    # Сохраняем файл в буфер
    buffer = BytesIO()
    wb.save(buffer)
    buffer.seek(0)

    # Создаем HTTP-ответ с файлом
    response = HttpResponse(
        buffer.getvalue(),
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = f'attachment; filename="комплектация_{record.id}.xlsx"'

    return response


def clear_products(request, pk):
    """Очистка продуктов - только для администраторов"""
    if not (request.user.is_authenticated and (request.user.is_staff or request.user.is_superuser)):
        messages.error(request, 'У вас нет прав для очистки комплектующих')
        return redirect('record_detail', pk=pk)
    
    record = get_object_or_404(Record, id=pk)
    RecordProduct.objects.filter(record=record).delete()
    record.products.clear()
    messages.success(request, "Комплектующие очищены.")
    return redirect('record_detail', pk=pk)


def add_products_to_record(request, pk):
    """Упрощенная функция добавления продуктов к записи"""
    record = get_object_or_404(Record, id=pk)
    
    if not (request.user.is_authenticated and (request.user.is_staff or request.user.is_superuser)):
        messages.error(request, 'У вас нет прав для добавления продуктов')
        return redirect('record_detail', pk=pk)

    # Инициализируем форму фильтрации
    filter_form = ProductFilterForm(request.GET)
    
    # Базовые queryset'ы
    products_qs = Product.objects.select_related('category').prefetch_related(
        'product_custom_fields__category_field'
    ).all()

    # Применяем фильтры к продуктам
    if filter_form.is_valid():
        filters = Q()
        
        # Поиск по названию
        if search := filter_form.cleaned_data.get('search'):
            filters &= Q(name__icontains=search)
        
        # Фильтры по категории
        if category := filter_form.cleaned_data.get('category'):
            filters &= Q(category=category)
        
        # Применяем основные фильтры
        if filters:
            products_qs = products_qs.filter(filters)
        
        # Фильтры по динамическим полям категории (из custom_fields)
        # Обрабатываем их отдельно для поддержки JSONField в SQLite
        for key, value in filter_form.cleaned_data.items():
            if key.startswith('filter_') and value:
                field_key = key.replace('filter_', '')
                # Используем специальный формат для JSONField в Django
                # Это работает и для SQLite, и для PostgreSQL
                products_qs = products_qs.filter(**{f'custom_fields__{field_key}': value})
    
    # Фильтры по ProductCustomField (индивидуальные характеристики)
    # Обрабатываем их отдельно, так как они не в форме
    for key, value in request.GET.items():
        if key.startswith('pcf_filter_') and value:
            try:
                category_field_id = int(key.replace('pcf_filter_', ''))
                from ..models import ProductCustomField
                # Фильтруем продукты, у которых есть ProductCustomField с таким значением
                product_ids = ProductCustomField.objects.filter(
                    category_field_id=category_field_id,
                    value=value
                ).values_list('product_id', flat=True)
                if product_ids:
                    products_qs = products_qs.filter(id__in=product_ids)
                else:
                    # Если нет продуктов с таким значением, возвращаем пустой queryset
                    products_qs = products_qs.none()
            except (ValueError, TypeError):
                pass

    # Простая сортировка по умолчанию (только для удобства отображения)
    products_qs = products_qs.order_by('name')

    # Загружаем текущие данные записи
    current_record_products = {
        rp.product_id: rp 
        for rp in RecordProduct.objects.filter(record=record).select_related('product')
    }

    # Обработка POST запроса
    if request.method == 'POST':
        selected_products = []
        
        # Обрабатываем продукты (включая плиты)
        for key, value in request.POST.items():
            if key.startswith('product_') and value == 'on':
                product_id = int(key.replace('product_', ''))
                quantity = request.POST.get(f'quantity_{product_id}', '1').strip()
                buyer = request.POST.get(f'buyer_{product_id}', 'Юра')
                custom_price = request.POST.get(f'custom_price_{product_id}', '').strip()
                
                try:
                    quantity = Decimal(quantity) if quantity else Decimal('1')
                    if quantity <= 0:
                        quantity = Decimal('1')
                except (InvalidOperation, ValueError):
                    quantity = Decimal('1')
                
                try:
                    custom_price = Decimal(custom_price) if custom_price else None
                except (InvalidOperation, ValueError):
                    custom_price = None
                
                selected_products.append({
                    'product_id': product_id,
                    'quantity': quantity,
                    'buyer': buyer,
                    'custom_price': custom_price
                })

        # Удаляем старые записи
        RecordProduct.objects.filter(record=record).delete()

        # Создаем новые записи для продуктов (включая плиты)
        if selected_products:
            product_ids = [p['product_id'] for p in selected_products]
            products_dict = {p.id: p for p in Product.objects.filter(id__in=product_ids)}
            
            record_products = [
                RecordProduct(
                    record=record,
                    product=products_dict[p['product_id']],
                    quantity=p['quantity'],
                    buyer=p['buyer'],
                    custom_price=p['custom_price']
                )
                for p in selected_products if p['product_id'] in products_dict
            ]
            RecordProduct.objects.bulk_create(record_products)
            record.products.set(products_dict.values())

        messages.success(request, "Комплектующие успешно сохранены!")
        return redirect('record_detail', pk=pk)

    # Подготавливаем данные для отображения (только товары, без плит)
    items_data = []
    
    # Добавляем только товары (максимум 30)
    for product in products_qs[:30]:
        rp = current_record_products.get(product.id)
        items_data.append({
            'type': 'product',
            'product': product,
            'is_selected': rp is not None,
            'quantity': rp.quantity if rp else 1,
            'buyer': rp.buyer if rp else 'Юра',
            'custom_price': rp.custom_price if rp else None,
        })

    # Получаем категории с их характеристиками для фильтров
    categories_with_fields = []
    selected_pcf_filters = {}  # Сохраняем выбранные значения для отображения
    from ..models import CategoryField, ProductCustomField
    for category in Category.objects.all():
        fields = category.category_fields.all()
        if fields.exists():
            # Получаем поля с ProductCustomField для фильтрации
            pcf_fields = []
            for field in fields:
                # Получаем уникальные значения для этого поля из ProductCustomField
                unique_values = ProductCustomField.objects.filter(
                    category_field=field,
                    product__in=Product.objects.all()  # Используем все продукты для получения всех значений
                ).exclude(value='').values_list('value', flat=True).distinct().order_by('value')
                
                if unique_values.exists():
                    field.unique_values = list(unique_values)
                    # Сохраняем выбранное значение для этого поля (используем строковый ключ)
                    selected_value = request.GET.get(f'pcf_filter_{field.id}', '')
                    selected_pcf_filters[str(field.id)] = selected_value
                    pcf_fields.append(field)
            
            categories_with_fields.append({
                'category': category,
                'fields': fields,
                'pcf_fields': pcf_fields
            })
    
    return render(request, 'add_products.html', {
        'record': record,
        'filter_form': filter_form,
        'items_data': items_data,
        'products_count': products_qs.count(),
        'categories_with_fields': categories_with_fields,
        'selected_pcf_filters': selected_pcf_filters,
    })


@login_required
def product_detail(request, pk):
    """Детальная информация и редактирование продукта"""
    product = get_object_or_404(Product.objects.select_related('category'), id=pk)
    
    category_fields_all = product.category.get_fields() if product.category else []
    existing_pcfs = {
        pcf.category_field_id: pcf
        for pcf in ProductCustomField.objects.filter(product=product).select_related('category_field')
    }

    if request.method == 'POST':
        # Простейшее редактирование: имя, категория, цена, изображение
        name = request.POST.get('name', '').strip()
        category_id = request.POST.get('category')
        our_price = request.POST.get('our_price', '').strip()
        source_url = request.POST.get('source_url', '').strip()
        image = request.FILES.get('image')
        remove_image = request.POST.get('remove_image')

        if name:
            product.name = name
        # Категория
        if category_id:
            try:
                product.category = Category.objects.get(id=category_id)
            except Category.DoesNotExist:
                product.category = None
        else:
            product.category = None
        # Цена: обновляем только если введено новое значение
        if our_price:
            try:
                product.our_price = Decimal(our_price)
            except (InvalidOperation, ValueError):
                messages.error(request, "Неверный формат цены")
                return redirect('product_detail', pk=product.id)

        # Ссылка (необязательно): можно очистить
        if source_url and '://' not in source_url:
            source_url = 'https://' + source_url
        product.source_url = source_url
        
        # Изображение: обновляем если загружено новое
        if image:
            product.image = image
            messages.success(request, "Изображение обновлено")
        
        # Удаление изображения
        if remove_image and product.image:
            product.image.delete()
            product.image = None
            messages.success(request, "Изображение удалено")

        product.save()
        
        # Информируем об изменении цены
        if our_price:
            messages.info(request, f"Цена обновлена: {product.our_price} ₽")

        # Сохраняем/удаляем индивидуальные характеристики по шаблонам категории
        if product.category:
            # обновляем существующие
            for field in product.category.get_fields():
                value = request.POST.get(f'pcf_value_{field.id}', '')
                delete_flag = request.POST.get(f'pcf_delete_{field.id}')
                pcf = existing_pcfs.get(field.id)
                if pcf:
                    if delete_flag:
                        # Удаляем сам объект ProductCustomField
                        pcf.delete()
                    else:
                        pcf.value = value.strip()
                        pcf.save()
                else:
                    if value.strip():
                        ProductCustomField.objects.create(
                            product=product,
                            category_field=field,
                            value=value.strip(),
                            order=0
                        )

            # добавление новых характеристик из выпадающих списков (может быть несколько)
            new_field_ids = request.POST.getlist('new_field[]')
            new_values = request.POST.getlist('new_value[]')
            new_template_names = request.POST.getlist('new_template_name[]')
            
            # Обрабатываем каждую тройку (шаблон, значение, название нового шаблона)
            from ..models import CategoryField
            for i, (new_field_id, new_value) in enumerate(zip(new_field_ids, new_values)):
                new_value = new_value.strip()
                new_template_name = new_template_names[i].strip() if i < len(new_template_names) else ''
                
                if not new_value:
                    continue
                
                category_field = None
                
                # Если выбрано "Создать новый шаблон"
                if new_field_id == '__new__' and new_template_name and product.category:
                    try:
                        # Создаем новый CategoryField для этой категории
                        category_field, created = CategoryField.objects.get_or_create(
                            category=product.category,
                            name=new_template_name,
                            defaults={
                                'field_type': 'text',
                                'required': False
                            }
                        )
                        if created:
                            messages.success(request, f'Создан новый шаблон: {new_template_name}')
                    except Exception as e:
                        messages.error(request, f'Ошибка создания шаблона: {str(e)}')
                        continue
                
                # Если выбран существующий шаблон
                elif new_field_id and new_field_id != '__new__':
                    try:
                        new_field_id_int = int(new_field_id)
                        category_field = product.category.category_fields.get(id=new_field_id_int)
                    except (ValueError, CategoryField.DoesNotExist, AttributeError):
                        continue
                
                # Создаем ProductCustomField если получили category_field
                if category_field:
                    # Перезагружаем existing_pcfs, так как могли добавить новые в цикле
                    current_pcf_ids = set(
                        ProductCustomField.objects.filter(product=product)
                        .values_list('category_field_id', flat=True)
                    )
                    # можно добавлять только если такого pcf еще нет
                    if category_field.id not in current_pcf_ids:
                        ProductCustomField.objects.create(
                            product=product,
                            category_field=category_field,
                            value=new_value,
                            order=0
                        )

        messages.success(request, "Изменения сохранены")
        return redirect('product_detail', pk=product.id)

    categories = Category.objects.all().order_by('name')
    category_fields = []
    available_fields = []
    if product.category:
        # Получаем только существующие ProductCustomField для этого продукта
        existing_pcf_list = ProductCustomField.objects.filter(product=product).select_related('category_field')
        
        # Формируем список только тех полей, для которых есть ProductCustomField
        for pcf in existing_pcf_list:
            category_fields.append({
                'field': pcf.category_field,
                'value': pcf.value
            })
        
        # Список доступных полей - все поля категории, для которых НЕТ ProductCustomField
        used_ids = {pcf.category_field_id for pcf in existing_pcf_list}
        available_fields = [f for f in product.category.get_fields() if f.id not in used_ids]

    return render(
        request,
        'product_detail.html',
        {
            'product': product,
            'categories': categories,
            'category_fields': category_fields,
            'available_fields': available_fields,
        }
    )


@login_required
def products_list(request):
    """Список всех продуктов (элементов)"""
    # Ограничиваем просмотр только для персонала, как в админке
    if not (request.user.is_staff or request.user.is_superuser):
        messages.error(request, 'Доступ к списку продуктов есть только у сотрудников.')
        return redirect('home')

    products = Product.objects.select_related('category').order_by('name')
    return render(request, 'products_list.html', {'products': products})


def get_mounting_types_by_category(category_name):
    """Возвращает доступные типы монтажа для категории"""
    if 'петл' in category_name.lower():
        return Product.objects.filter(
            category__name__icontains='петл'
        ).exclude(mounting_type__isnull=True).exclude(mounting_type='').values_list(
            'mounting_type', flat=True
        ).distinct()
    elif 'направля' in category_name.lower():
        return Product.objects.filter(
            category__name__icontains='направля'
        ).exclude(mounting_type__isnull=True).exclude(mounting_type='').values_list(
            'mounting_type', flat=True
        ).distinct()
    return []


# Устаревшие функции (для обратной совместимости)
def get_excel_data(request, pk):
    return JsonResponse({'data': [], 'success': False, 'error': 'Function deprecated'})


def save_excel_data(request, pk):
    return JsonResponse({'success': False, 'error': 'Function deprecated'})


def download_excel_file(request, pk):
    return export_products(request, pk)
