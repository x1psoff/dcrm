from django.shortcuts import render, redirect, get_object_or_404
from django.contrib.auth import authenticate, login, logout
from django.contrib import messages
from .forms import SignUpForm, AddRecordForm, RecordProductForm, HingeFilterForm, RunnerFilterForm, ProductFilterForm, UnplannedExpenseForm
from django.contrib.auth.decorators import login_required
from .models import RecordProduct, Record, UploadedFile, Product, Brand, Category, UnplannedExpense, Plita, RecordPlita, Designer
import pandas as pd
from django.conf import settings
import os
from django.http import HttpResponse,FileResponse, JsonResponse
from django.db import models
import csv
from openpyxl import load_workbook
from io import BytesIO
from decimal import Decimal, InvalidOperation
import json
import re
from django.db.models import Sum, Count, Avg
from collections import defaultdict
import calendar
from datetime import datetime, timedelta
@login_required
def unplanned_expenses_list(request):
    """Список всех непланируемых расходов"""
    expenses = UnplannedExpense.objects.all()
    total = sum(expense.price for expense in expenses)

    return render(request, 'unplanned_expenses/list.html', {
        'expenses': expenses,
        'total': total,
        'form': UnplannedExpenseForm()  # Форма для добавления новых расходов
    })


@login_required
def add_unplanned_expense(request, pk):
    """Добавление непланируемого расхода к записи"""
    record = get_object_or_404(Record, id=pk)

    if request.method == 'POST':
        item = request.POST.get('item')
        price = request.POST.get('price')
        spent_by = request.POST.get('spent_by', 'Юра')  # По умолчанию Юра

        if item and price:
            UnplannedExpense.objects.create(
                record=record,
                item=item,
                price=price,
                spent_by=spent_by
            )
            messages.success(request, "Расход успешно добавлен!")
        else:
            messages.error(request, "Пожалуйста, заполните все обязательные поля.")

    return redirect('record_detail', pk=pk)


@login_required
def edit_unplanned_expense(request, pk):
    """Редактирование существующего расхода"""
    expense = get_object_or_404(UnplannedExpense, id=pk)

    if request.method == 'POST':
        form = UnplannedExpenseForm(request.POST, instance=expense)
        if form.is_valid():
            form.save()
            messages.success(request, "Расход успешно обновлен!")
            return redirect('unplanned_expenses_list')
    else:
        form = UnplannedExpenseForm(instance=expense)

    return render(request, 'unplanned_expenses/edit.html', {
        'form': form,
        'expense': expense
    })


@login_required
def delete_unplanned_expense(request, pk):
    """Удаление непланируемого расхода"""
    expense = get_object_or_404(UnplannedExpense, id=pk)
    record_id = expense.record.id
    expense.delete()
    messages.success(request, "Расход успешно удален!")
    return redirect('record_detail', pk=record_id)


def export_products(request, pk):
    record = get_object_or_404(Record, id=pk)

    # Загружаем шаблон Excel
    template_path = os.path.join(settings.MEDIA_ROOT, 'templates', 'комплектация Кухни.xlsx')
    wb = load_workbook(template_path)
    ws = wb.active

    # Сопоставление типов петель со строками в Excel
    petli_mapping = {
        'накладная': 28,
        'пружинная': 29,
        'смежная': 30,
        'вкладная': 31,
        'фальш-планка': 32,
        '155гр.': 33,
        'для алюм рамок': 34
    }

    # Сопоставление типов направляющих со строками в Excel
    naprav_mapping = {
        'частичного выдвижения': 35,
        'скрыт.мон.дов': 36,
        'скрыт.мон.типон': 37,
        'полного выдвижения': 44
    }

    # Обрабатываем продукты записи
    for product in record.products.all():
        try:
            record_product = RecordProduct.objects.get(record=record, product=product)
            buyer = record_product.buyer
            quantity = record_product.quantity

            # Определяем категорию продукта
            category_name = product.category.name.lower() if product.category else ""

            # Обрабатываем петли
            if 'петл' in category_name and product.mounting_type in petli_mapping:
                row = petli_mapping[product.mounting_type]

                # Записываем количество
                ws[f'D{row}'] = quantity

                # Записываем название продукта
                ws[f'E{row}'] = product.name

                # Записываем цену в зависимости от покупателя
                if buyer == 'Юра':
                    ws[f'G{row}'] = product.our_price * quantity
                else:
                    ws[f'H{row}'] = product.our_price * quantity

            # Обрабатываем направляющие
            elif 'направля' in category_name:
                # Определяем тип направляющих
                naprav_type = None
                if product.runner_size:
                    naprav_type = product.runner_size
                elif product.response_type:
                    naprav_type = product.response_type

                if naprav_type in naprav_mapping:
                    row = naprav_mapping[naprav_type]

                    # Записываем количество
                    ws[f'D{row}'] = quantity

                    # Записываем название продукта
                    ws[f'E{row}'] = product.name

                    # Записываем цену в зависимости от покупателя
                    if buyer == 'Юра':
                        ws[f'G{row}'] = product.our_price * quantity
                    else:
                        ws[f'H{row}'] = product.our_price * quantity
                else:
                    # Для других типов направляющих используем строку 48
                    row = 48

                    # Записываем тип в столбец B
                    ws[f'B{row}'] = naprav_type

                    # Записываем количество
                    ws[f'D{row}'] = quantity

                    # Записываем название продукта
                    ws[f'E{row}'] = product.name

                    # Записываем цену в зависимости от покупателя
                    if buyer == 'Юра':
                        ws[f'G{row}'] = product.our_price * quantity
                    else:
                        ws[f'H{row}'] = product.our_price * quantity

            # Обрабатываем другие продукты
            else:
                # Находим первую свободную строку после 48
                row = 48
                while ws[f'B{row}'].value is not None:
                    row += 1

                # Записываем тип в столбец B
                ws[f'B{row}'] = category_name

                # Записываем количество
                ws[f'D{row}'] = quantity

                # Записываем название продукта
                ws[f'E{row}'] = product.name

                # Записываем цену в зависимости от покупателя
                if buyer == 'Юра':
                    ws[f'G{row}'] = product.our_price * quantity
                else:
                    ws[f'H{row}'] = product.our_price * quantity

        except RecordProduct.DoesNotExist:
            continue

    # Сохраняем файл в буфер
    buffer = BytesIO()
    wb.save(buffer)
    buffer.seek(0)

    # Создаем HTTP-ответ с файлом
    response = HttpResponse(
        buffer.getvalue(),
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = f'attachment; filename="комплектация_{record.id}.xlsx"'

    return response

def clear_products(request, pk):
    if request.user.is_authenticated:
        record = get_object_or_404(Record, id=pk)
        # Удаляем все товарные позиции
        record.products.clear()
        # Удаляем все плиты (через промежуточную таблицу с количеством)
        RecordPlita.objects.filter(record=record).delete()
        # Также сбросим M2M связь на всякий случай
        record.plitas.clear()
        messages.success(request, "Комплектующие и плиты очищены.")
        return redirect('record_detail', pk=pk)
    else:
        messages.error(request, "Вы должны быть авторизованы.")
        return redirect('home')





def home(request):
    records = Record.objects.all()
    
    # Статистика по статусам
    status_stats = {
        'otrisovka': records.filter(status='otrisovka').count(),
        'zhdem_material': records.filter(status='zhdem_material').count(),
        'priekhal_v_ceh': records.filter(status='priekhal_v_ceh').count(),
        'na_raspile': records.filter(status='na_raspile').count(),
        'zakaz_gotov': records.filter(status='zakaz_gotov').count(),
    }
    
    # Общая сумма по договорам
    total_contract_amount = sum(record.contract_amount for record in records if record.contract_amount)
    
    if request.method == 'POST':
        username = request.POST['username']
        password = request.POST['password']
        user = authenticate(request, username=username, password=password)
        if user is not None:
            login(request, user)
            messages.success(request, "You Have Been Logged In!")
            return redirect('home')
        else:
            messages.success(request, "There Was An Error Logging In, Please Try Again...")
            return redirect('home')
    else:
        return render(request, 'home.html', {
            'records': records,
            'status_stats': status_stats,
            'total_contract_amount': total_contract_amount
        })

def logout_user(request):
    logout(request)
    messages.success(request, "You Have Been Logged Out...")
    return redirect('home')


def register_user(request):
    if request.method == 'POST':
        form = SignUpForm(request.POST)
        if form.is_valid():
            form.save()
            username = form.cleaned_data['username']
            password = form.cleaned_data['password1']
            user = authenticate(username=username, password=password)
            login(request, user)
            messages.success(request, "You Have Successfully Registered! Welcome!")
            return redirect('home')
    else:
        form = SignUpForm()
    return render(request, 'register.html', {'form': form})


def customer_record(request, pk):
    if request.user.is_authenticated:
        customer_record = Record.objects.get(id=pk)
        return render(request, 'record.html', {'customer_record': customer_record})
    else:
        messages.success(request, "You Must Be Logged In To View That Page...")
        return redirect('home')


def delete_record(request, pk):
    if request.user.is_authenticated:
        delete_it = Record.objects.get(id=pk)
        delete_it.delete()
        messages.success(request, "Record Deleted Successfully...")
        return redirect('home')
    else:
        messages.success(request, "You Must Be Logged In To Do That...")
        return redirect('home')


def add_record(request):
    if request.method == 'POST':
        form = AddRecordForm(request.POST, request.FILES)
        if form.is_valid():
            record = form.save()
            if 'file' in request.FILES:
                UploadedFile.objects.create(record=record, file=request.FILES['file'])
            messages.success(request, "Запись успешно добавлена!")
            return redirect('home')
    else:
        form = AddRecordForm()
    return render(request, 'add_record.html', {'form': form})


def add_products_to_record(request, pk):
    record = get_object_or_404(Record, id=pk)

    # Используем общую форму фильтрации
    filter_form = ProductFilterForm(request.GET)
    products_queryset = Product.objects.all()
    plitas_queryset = Plita.objects.all()

    if filter_form.is_valid():
        if filter_form.cleaned_data.get('name'):
            products_queryset = products_queryset.filter(
                name__icontains=filter_form.cleaned_data['name']
            )
        if filter_form.cleaned_data.get('category'):
            products_queryset = products_queryset.filter(
                category=filter_form.cleaned_data['category']
            )
        if filter_form.cleaned_data.get('brand'):
            products_queryset = products_queryset.filter(
                brand=filter_form.cleaned_data['brand']
            )
        if filter_form.cleaned_data.get('hinge_angle'):
            products_queryset = products_queryset.filter(
                hinge_angle=filter_form.cleaned_data['hinge_angle']
            )
        if filter_form.cleaned_data.get('hinge_closing_type'):
            products_queryset = products_queryset.filter(
                hinge_closing_type=filter_form.cleaned_data['hinge_closing_type']
            )
        if filter_form.cleaned_data.get('response_type'):
            products_queryset = products_queryset.filter(
                response_type=filter_form.cleaned_data['response_type']
            )
        if filter_form.cleaned_data.get('runner_size'):
            products_queryset = products_queryset.filter(
                runner_size=filter_form.cleaned_data['runner_size']
            )
        # Общий фильтр по типу монтажа, если задан через форму
        if filter_form.cleaned_data.get('mounting_type'):
            products_queryset = products_queryset.filter(
                mounting_type=filter_form.cleaned_data['mounting_type']
            )

    # Обрабатываем новые параметры фильтрации для mounting_type
    hinge_mounting_type = request.GET.get('hinge_mounting_type')
    mounting_type_param = request.GET.get('mounting_type')  # общий select name="mounting_type"
    # Фильтры плит
    plita_material = request.GET.get('plita_material')
    plita_thickness = request.GET.get('plita_thickness')

    if hinge_mounting_type:
        products_queryset = products_queryset.filter(mounting_type=hinge_mounting_type)

    # Применяем общий параметр mounting_type, если он присутствует (нормализуем пробелы/подчеркивания)
    if mounting_type_param:
        raw = mounting_type_param.strip()
        # Нормализуем возможную латинскую 'C' в кириллическую 'С' и наоборот для устойчивости
        swap_c = str.maketrans({'C': 'С', 'c': 'с'})
        swap_ru = str.maketrans({'С': 'C', 'с': 'c'})
        raw_cyr = raw.translate(swap_c)
        raw_lat = raw.translate(swap_ru)
        equivalents = list({
            raw,
            raw.replace('_', ' '),
            raw.replace(' ', '_'),
            raw_cyr,
            raw_cyr.replace('_', ' '),
            raw_cyr.replace(' ', '_'),
            raw_lat,
            raw_lat.replace('_', ' '),
            raw_lat.replace(' ', '_'),
        })
        # Нормализуем регистр и пробелы/подчёркивания в БД через icontains/iregex
        q_objects = (
            models.Q(mounting_type__in=equivalents) |
            models.Q(mounting_type__iexact=raw) |
            models.Q(mounting_type__iexact=raw.replace('_',' ')) |
            models.Q(mounting_type__iexact=raw.replace(' ','_')) |
            models.Q(mounting_type__iexact=raw_cyr) |
            models.Q(mounting_type__iexact=raw_lat) |
            models.Q(mounting_type__icontains=raw) |
            models.Q(mounting_type__icontains=raw_cyr) |
            models.Q(mounting_type__icontains=raw_lat)
        )
        products_queryset = products_queryset.filter(q_objects)

    if plita_material:
        plitas_queryset = plitas_queryset.filter(material=plita_material)
    if plita_thickness:
        try:
            plitas_queryset = plitas_queryset.filter(thickness=plita_thickness)
        except Exception:
            pass

    # Добавляем сортировку
    sort_by = request.GET.get('sort_by', 'name')
    sort_order = request.GET.get('sort_order', 'asc')

    # Определяем поле для сортировки
    if sort_by == 'price':
        sort_field = 'our_price'
    elif sort_by == 'brand':
        sort_field = 'brand__name'
    elif sort_by == 'category':
        sort_field = 'category__name'
    else:  # по умолчанию сортируем по названию
        sort_field = 'name'

    # Определяем направление сортировки
    if sort_order == 'desc':
        sort_field = '-' + sort_field

    # Применяем сортировку
    products_queryset = products_queryset.order_by(sort_field)

    # Для GET запроса подготавливаем данные о текущих выборах
    current_products = record.products.all()
    current_buyers = {}
    current_quantities = {}
    current_custom_prices = {}
    current_plita_quantities = {}
    for rp in RecordProduct.objects.filter(record=record):
        current_buyers[rp.product_id] = rp.buyer
        current_quantities[rp.product_id] = rp.quantity
        current_custom_prices[rp.product_id] = rp.custom_price
    for rpl in RecordPlita.objects.filter(record=record):
        current_plita_quantities[rpl.plita_id] = rpl.quantity

    # Форма добавления продуктов
    if request.method == 'POST':
        # Получаем список выбранных продуктов и их данные
        selected_products_data = []
        selected_plitas_data = []
        
        for key in request.POST.keys():
            if key.startswith('product_'):
                product_id = key.replace('product_', '')
                
                # Validate that product_id is a valid integer
                try:
                    product_id_int = int(product_id)
                except (ValueError, TypeError):
                    continue  # Skip invalid product IDs
                
                buyer = request.POST.get(f'buyer_{product_id}', 'Юра')
                quantity = request.POST.get(f'quantity_{product_id}', 1)
                custom_price_raw = request.POST.get(f'custom_price_{product_id}', '').strip()

                try:
                    quantity = int(quantity)
                except (ValueError, TypeError):
                    quantity = 1

                custom_price = None
                if custom_price_raw != '':
                    try:
                        custom_price = Decimal(custom_price_raw)
                    except (InvalidOperation, ValueError):
                        custom_price = None

                selected_products_data.append({
                    'product_id': product_id_int,  # Use the validated integer
                    'buyer': buyer,
                    'quantity': quantity,
                    'custom_price': custom_price
                })

            elif key.startswith('plita_'):
                plita_id = key.replace('plita_', '')
                
                # Validate that plita_id is a valid integer
                try:
                    plita_id_int = int(plita_id)
                except (ValueError, TypeError):
                    continue  # Skip invalid plita IDs
                
                qty_raw = request.POST.get(f'plita_quantity_{plita_id}', '1').strip()
                try:
                    qty = Decimal(qty_raw)
                    if qty <= 0:
                        raise InvalidOperation
                except Exception:
                    qty = Decimal('1')
                selected_plitas_data.append({
                    'plita_id': plita_id_int,  # Use the validated integer
                    'quantity': qty,
                })

        # Удаляем все RecordProduct и RecordPlita для этой записи
        RecordProduct.objects.filter(record=record).delete()
        RecordPlita.objects.filter(record=record).delete()

        # Создаем новые записи для выбранных продуктов
        for product_data in selected_products_data:
            try:
                product = Product.objects.get(id=product_data['product_id'])
                RecordProduct.objects.create(
                    record=record,
                    product=product,
                    buyer=product_data['buyer'],
                    quantity=product_data['quantity'],
                    custom_price=product_data['custom_price']
                )
            except Product.DoesNotExist:
                continue  # Skip if product doesn't exist

        # Создаем новые записи для выбранных плит
        for plita_data in selected_plitas_data:
            try:
                plita = Plita.objects.get(id=plita_data['plita_id'])
                RecordPlita.objects.create(
                    record=record,
                    plita=plita,
                    quantity=plita_data['quantity']
                )
            except Plita.DoesNotExist:
                continue  # Skip if plita doesn't exist

        # Обновляем связь ManyToMany
        record.products.set(Product.objects.filter(
            id__in=[p['product_id'] for p in selected_products_data]
        ))

        messages.success(request, "Комплектующие добавлены!")
        return redirect('record_detail', pk=pk)

    # Подготавливаем данные для отображения
    products_data = []
    plitas_data = []
    for product in products_queryset:
        products_data.append({
            'product': product,
            'is_checked': product in current_products,
            'buyer': current_buyers.get(product.id, 'Юра'),
            'quantity': current_quantities.get(product.id, 1),
            'custom_price': current_custom_prices.get(product.id)
        })
    for plita in plitas_queryset:
        plitas_data.append({
            'plita': plita,
            'is_checked': plita.id in current_plita_quantities,
            'quantity': current_plita_quantities.get(plita.id, 1)
        })

    return render(request, 'add_products.html', {
        'products_data': products_data,
        'filter_form': filter_form,
        'record': record,
        'plita_material_choices': Plita.objects.values_list('material', flat=True).distinct(),
        'plita_thickness_choices': Plita.objects.values_list('thickness', flat=True).distinct(),
        'selected_plita_material': plita_material or '',
        'selected_plita_thickness': plita_thickness or '',
        'plitas_data': plitas_data,
    })

def update_record(request, pk):
    current_record = get_object_or_404(Record, id=pk)
    if request.method == 'POST':
        form = AddRecordForm(request.POST, request.FILES, instance=current_record)
        if form.is_valid():
            record = form.save()
            if 'file' in request.FILES:
                UploadedFile.objects.create(record=record, file=request.FILES['file'])
            messages.success(request, "Запись успешно обновлена!")
            return redirect('record_detail', pk=record.id)  # Исправлено с 'record' на 'record_detail'
    else:
        form = AddRecordForm(instance=current_record)
    return render(request, 'update_record.html', {
        'form': form,
        'current_record': current_record
    })


@login_required
def set_margin_flags(request, pk):
    record = get_object_or_404(Record, id=pk)
    if request.method == 'POST':
        record.margin_yura = bool(request.POST.get('margin_yura'))
        record.margin_oleg = bool(request.POST.get('margin_oleg'))
        record.save(update_fields=['margin_yura', 'margin_oleg'])
        messages.success(request, 'Флаги распределения моржи сохранены')
    return redirect('record_detail', pk=pk)


def record_detail(request, pk):
    customer_record = get_object_or_404(Record, id=pk)
    categories = Category.objects.all()
    brands = Brand.objects.all()

    # Фильтрация продуктов по категории из GET-параметра
    product_filter = request.GET.get('category', None)
    products = customer_record.products.select_related('category', 'brand', 'store').all()
    if product_filter:
        products = products.filter(category__name=product_filter)

    # Расчет сумм с учетом своей цены и количества
    rps = RecordProduct.objects.filter(record=customer_record).select_related('product')
    rp_map = {rp.product_id: rp for rp in rps}
    total_products = sum(
        (rp.custom_price if rp.custom_price is not None else rp.product.our_price) * rp.quantity
        for rp in rps
    )

    # Плиты для записи (отображение)
    record_plitas = RecordPlita.objects.filter(record=customer_record).select_related('plita')
    # Сумма по плитам
    try:
        plitas_total = sum((rpl.plita.price_per_square_meter or 0) * (rpl.quantity or 0) for rpl in record_plitas)
    except Exception:
        plitas_total = 0
    try:
        total_components = total_products + plitas_total
    except Exception:
        total_components = total_products

    # Создаем единую таблицу всех компонентов мебели
    unified_components = []
    
    # Добавляем продукты (направляющие, петли и другие)
    for product in products:
        category_name = product.category.name if product.category else 'Без категории'
        rp = rp_map.get(product.id)
        unit_price = (rp.custom_price if (rp and rp.custom_price is not None) else product.our_price)
        quantity = rp.quantity if rp else 1
        try:
            line_total = unit_price * quantity
        except Exception:
            line_total = unit_price
        
        # Определяем тип компонента для группировки
        component_type = 'Другие'
        if 'петл' in category_name.lower():
            component_type = 'Петли'
        elif 'направля' in category_name.lower():
            component_type = 'Направляющие'
        
        unified_components.append({
            'type': component_type,
            'name': product.name,
            'category': category_name,
            'brand': product.brand.name if product.brand else '',
            'specifications': {
                'hinge_angle': product.hinge_angle,
                'hinge_closing_type': product.hinge_closing_type,
                'runner_size': product.runner_size,
                'response_type': product.response_type,
                'mounting_type': product.mounting_type,
            },
            'unit_price': unit_price,
            'quantity': quantity,
            'line_total': line_total,
            'buyer': rp.buyer if rp else 'Юра',
            'is_product': True,
            'product_id': product.id
        })
    
    # Добавляем плиты
    for rp in record_plitas:
        unified_components.append({
            'type': 'Плиты',
            'name': rp.plita.name,
            'category': 'Плиты',
            'brand': '',
            'specifications': {
                'material': rp.plita.material,
                'thickness': f"{rp.plita.thickness}мм",
                'color': rp.plita.color,
            },
            'unit_price': rp.plita.price_per_square_meter,
            'quantity': rp.quantity,
            'line_total': rp.plita.price_per_square_meter * rp.quantity,
            'buyer': 'Юра',  # Плиты всегда покупает Юра
            'is_product': False,
            'plita_id': rp.plita.id
        })
    
    # Сортируем компоненты по типу, затем по названию
    unified_components.sort(key=lambda x: (x['type'], x['name']))
    
    # Группировка продуктов по категориям вместе с RecordProduct (для обратной совместимости)
    products_by_category = {}
    category_counts = {}
    for product in products:
        category_name = product.category.name if product.category else 'Без категории'
        rp = rp_map.get(product.id)
        unit_price = (rp.custom_price if (rp and rp.custom_price is not None) else product.our_price)
        quantity = rp.quantity if rp else 1
        try:
            line_total = unit_price * quantity
        except Exception:
            line_total = unit_price
        products_by_category.setdefault(category_name, []).append({
            'product': product,
            'rp': rp,
            'unit_price': unit_price,
            'quantity': quantity,
            'line_total': line_total,
        })
        category_counts[category_name] = category_counts.get(category_name, 0) + 1

    # Сортировка категорий по имени
    products_by_category = dict(sorted(products_by_category.items(), key=lambda kv: kv[0].lower()))
    category_counts = dict(sorted(category_counts.items(), key=lambda kv: kv[0].lower()))

    # Сводные счётчики по основным группам
    hinge_count = sum(len(items) for name, items in products_by_category.items() if 'петл' in name.lower())
    runner_count = sum(len(items) for name, items in products_by_category.items() if 'направля' in name.lower())
    plita_count = record_plitas.count()
    total_expenses = sum(expense.price for expense in customer_record.unplanned_expenses.all())

    # Расчет зарплаты проектировщика
    designer_salary = 0
    designer_info = ''
    if customer_record.designer:
        d = customer_record.designer
        
        # Определяем метод расчета по полю method или по наличию данных
        method_name = d.method.name.lower() if d.method else ''
        
        # Метод 1: процент от суммы договора
        if ('процент' in method_name or (not d.method and d.percentage)) and d.percentage and customer_record.contract_amount:
            try:
                designer_salary = (customer_record.contract_amount * d.percentage) / 100
                designer_info = f"{d.name} {d.surname} — {d.percentage}% от договора"
            except Exception:
                designer_salary = 0
        # Метод 2: ставка за м² по данным CSV (process_csv)
        elif ('м²' in method_name or 'метр' in method_name or (not d.method and d.rate_per_square_meter)) and d.rate_per_square_meter:
            try:
                uploaded_files = customer_record.files.all()
                sum_area = 0
                for uploaded_file in uploaded_files:
                    file_path = os.path.join(settings.MEDIA_ROOT, uploaded_file.file.name)
                    try:
                        df = pd.read_csv(
                            file_path,
                            sep=';',
                            encoding='cp1251',
                            header=None,
                            engine='python',
                            on_bad_lines='warn'
                        )
                        if df.shape[1] >= 5:
                            for col in [2, 3, 4]:
                                df[col] = pd.to_numeric(df[col], errors='coerce')
                            df[2] = df[2].round() / 1000
                            df[3] = df[3].round() / 1000
                            filtered_df = df[df[4].isin([16, 18])].copy()
                            filtered_df['area'] = (filtered_df[2] * filtered_df[3]).round(3)
                            sum_area += filtered_df['area'].sum()
                    except Exception:
                        continue
                designer_salary = (d.rate_per_square_meter or 0) * sum_area
                designer_info = f"{d.name} {d.surname} — {d.rate_per_square_meter} ₽/м² ({sum_area:.2f} м²)"
            except Exception:
                designer_salary = 0
        # Метод 3: погонный метр — ручной ввод погонных метров, умножаем на ставку за м² (rate_per_square_meter)
        elif 'погон' in method_name and customer_record.designer_manual_salary is not None:
            try:
                meters = Decimal(str(customer_record.designer_manual_salary))
            except Exception:
                meters = Decimal('0')
            rate = d.rate_per_square_meter or Decimal('0')
            designer_salary = rate * meters
            designer_info = f"{d.name} {d.surname} — {rate} ₽/м × {meters} м"

    total_amount = total_components + total_expenses + designer_salary

    # Распределение моржи
    # Моржа = total_amount - (себестоимость). Если себестоимость = total_products + total_expenses + designer_salary,
    # то моржа будет 0. Введём моржу как сумма по договору (contract_amount) - total_amount, если contract_amount задана
    margin_total = Decimal('0')
    if customer_record.contract_amount:
        try:
            margin_total = Decimal(customer_record.contract_amount) - Decimal(total_amount)
        except Exception:
            margin_total = Decimal('0')

    # Учёт индивидуальных расходов по людям (комплектующие + неплановые)
    # 1) Неплановые
    unplanned_spent = {
        'Юра': sum(e.price for e in customer_record.unplanned_expenses.filter(spent_by='Юра')),
        'Олег': sum(e.price for e in customer_record.unplanned_expenses.filter(spent_by='Олег')),
    }
    # 2) Комплектующие (RecordProduct.buyer)
    buyer_product_spent = {'Юра': Decimal('0'), 'Олег': Decimal('0')}
    for rp in RecordProduct.objects.filter(record=customer_record).select_related('product'):
        unit_price = rp.custom_price if rp.custom_price is not None else rp.product.our_price
        try:
            line_total = Decimal(unit_price) * Decimal(rp.quantity)
        except Exception:
            line_total = Decimal('0')
        if rp.buyer in buyer_product_spent:
            buyer_product_spent[rp.buyer] += line_total

    spent_by_totals = {
        'Юра': unplanned_spent['Юра'] + buyer_product_spent['Юра'],
        'Олег': unplanned_spent['Олег'] + buyer_product_spent['Олег'],
    }

    # Базовое распределение по чекбоксам
    recipients = []
    if customer_record.margin_yura:
        recipients.append('Юра')
    if customer_record.margin_oleg:
        recipients.append('Олег')
    if not recipients:
        recipients = ['Юра', 'Олег']

    per_head = (margin_total / len(recipients)) if recipients else Decimal('0')
    # Коррекция по правилу: вычесть траты одного из моржи другого
    margin_distribution = {}
    if 'Юра' in recipients:
        margin_distribution['Юра'] = per_head - Decimal(spent_by_totals.get('Олег', 0))
    if 'Олег' in recipients:
        margin_distribution['Олег'] = per_head - Decimal(spent_by_totals.get('Юра', 0))

    margin_yura_value = margin_distribution.get('Юра', Decimal('0'))
    margin_oleg_value = margin_distribution.get('Олег', Decimal('0'))
    spent_yura = spent_by_totals.get('Юра', Decimal('0'))
    spent_oleg = spent_by_totals.get('Олег', Decimal('0'))

    context = {
        'customer_record': customer_record,
        'categories': categories,
        'brands': brands,
        'products': products,
        'products_by_category': products_by_category,
        'category_counts': category_counts,
        'record_plitas': record_plitas,
        'unified_components': unified_components,
        'hinge_count': hinge_count,
        'runner_count': runner_count,
        'plita_count': plita_count,
        'total_products': total_products,
        'plitas_total': plitas_total,
        'total_components': total_components,
        'total_expenses': total_expenses,
        'designer_salary': designer_salary,
        'designer_info': designer_info,
        'total_amount': total_amount,
        'margin_total': margin_total,
        'margin_distribution': margin_distribution,
        'margin_yura_value': margin_yura_value,
        'margin_oleg_value': margin_oleg_value,
        'spent_yura': spent_yura,
        'spent_oleg': spent_oleg,
    }
    return render(request, 'record.html', context)

@login_required
def set_designer_manual_salary(request, pk):
    record = get_object_or_404(Record, id=pk)
    if request.method == 'POST':
        value = request.POST.get('designer_manual_salary', '').strip()
        try:
            record.designer_manual_salary = Decimal(value)
            record.save(update_fields=['designer_manual_salary'])
            messages.success(request, 'Сумма проектировщика сохранена')
        except Exception:
            messages.error(request, 'Неверное значение суммы')
    return redirect('record_detail', pk=pk)

def delete_file(request, file_id):
    file = get_object_or_404(UploadedFile, id=file_id)
    record_id = file.record.id
    file.delete()
    messages.success(request, "Файл успешно удален")
    return redirect('update_record', pk=record_id)


def process_csv(request):
    record_id_raw = request.GET.get('record_id')
    if not record_id_raw:
        return render(request, 'error.html', {'error': 'Не указан ID записи'})

    # Надёжно приводим record_id к целому числу, вычищая возможные артефакты шаблона
    record_id_clean = str(record_id_raw).strip()
    # Удаляем потенциальные остатки шаблонов вида '{{ ... }}'
    record_id_clean = record_id_clean.replace('{', '').replace('}', '')
    record_id_clean = record_id_clean.replace('%', '')
    record_id_clean = record_id_clean.replace('customer_record.id', '')
    record_id_clean = record_id_clean.strip()
    # Оставляем только цифры
    record_id_digits = ''.join(ch for ch in record_id_clean if ch.isdigit())
    try:
        record_id = int(record_id_digits)
    except (TypeError, ValueError):
        return render(request, 'error.html', {'error': f"Некорректный ID записи: {record_id_raw}"})

    record = get_object_or_404(Record, id=record_id)
    uploaded_files = record.files.all()

    files_data = []

    for uploaded_file in uploaded_files:
        file_info = {
            'file_name': os.path.basename(uploaded_file.file.name),
            'success': False
        }

        file_path = os.path.join(settings.MEDIA_ROOT, uploaded_file.file.name)

        try:
            # Чтение файла с фиксированной кодировкой cp1251
            df = pd.read_csv(
                file_path,
                sep=';',  # явно указываем разделитель
                encoding='cp1251',
                header=None,
                engine='python',
                on_bad_lines='warn'
            )

            # Проверка структуры (минимум 5 столбцов)
            if df.shape[1] < 5:
                raise ValueError("Файл должен содержать минимум 5 столбцов")

            # Обработка числовых колонок
            for col in [2, 3, 4]:  # колонки 3,4,5 (индексы 2,3,4)
                df[col] = pd.to_numeric(df[col], errors='coerce')

            # Фильтрация и расчеты
            df[2] = df[2].round() / 1000  # мм -> метры
            df[3] = df[3].round() / 1000
            filtered_df = df[df[4].isin([16, 18])].copy()
            filtered_df['area'] = (filtered_df[2] * filtered_df[3]).round(3)

            file_info.update({
                'data': filtered_df.to_dict('records'),
                'sum_16': filtered_df[filtered_df[4] == 16]['area'].sum().round(1),
                'sum_18': filtered_df[filtered_df[4] == 18]['area'].sum().round(1),
                'row_count': len(filtered_df),
                'success': True
            })

        except Exception as e:
            file_info['error'] = f"Ошибка обработки: {str(e)}"

        files_data.append(file_info)

    context = {
        'record': record,
        'files_data': files_data,
        'success_count': sum(1 for f in files_data if f['success'])
    }
    return render(request, 'process_csv.html', context)


def process_csv_by_pk(request, pk):
    """Альтернативный маршрут: принимает ID записи в URL без query-параметров."""
    record = get_object_or_404(Record, id=pk)

    uploaded_files = record.files.all()
    files_data = []
    total_area = 0  # Общая площадь из всех файлов

    for uploaded_file in uploaded_files:
        file_info = {
            'file_name': os.path.basename(uploaded_file.file.name),
            'success': False
        }

        file_path = os.path.join(settings.MEDIA_ROOT, uploaded_file.file.name)

        try:
            df = pd.read_csv(
                file_path,
                sep=';',
                encoding='cp1251',
                header=None,
                engine='python',
                on_bad_lines='warn'
            )

            if df.shape[1] < 5:
                raise ValueError("Файл должен содержать минимум 5 столбцов")

            for col in [2, 3, 4]:
                df[col] = pd.to_numeric(df[col], errors='coerce')

            df[2] = df[2].round() / 1000
            df[3] = df[3].round() / 1000
            filtered_df = df[df[4].isin([16, 18])].copy()
            filtered_df['area'] = (filtered_df[2] * filtered_df[3]).round(3)

            file_area = filtered_df['area'].sum()
            total_area += file_area

            file_info.update({
                'data': filtered_df.to_dict('records'),
                'sum_16': filtered_df[filtered_df[4] == 16]['area'].sum().round(1),
                'sum_18': filtered_df[filtered_df[4] == 18]['area'].sum().round(1),
                'row_count': len(filtered_df),
                'success': True
            })

        except Exception as e:
            file_info['error'] = f"Ошибка обработки: {str(e)}"

        files_data.append(file_info)

    # Расчет зарплаты проектировщика
    designer_salary = 0
    designer_info = ''
    if record.designer and record.designer.rate_per_square_meter:
        d = record.designer
        try:
            total_area_dec = Decimal(str(total_area))
        except Exception:
            total_area_dec = Decimal('0')
        base_rate = d.rate_per_square_meter or Decimal('0')
        designer_salary = base_rate * total_area_dec
        designer_info = f"{d.name} {d.surname} — {d.rate_per_square_meter} ₽/м²"
    elif record.designer and record.designer.percentage and record.contract_amount:
        d = record.designer
        designer_salary = (record.contract_amount * d.percentage) / 100
        designer_info = f"{d.name} {d.surname} — {d.percentage}% от договора"

    context = {
        'record': record,
        'files_data': files_data,
        'success_count': sum(1 for f in files_data if f['success']),
        'total_area': round(total_area, 2),
        'designer_salary': designer_salary,
        'designer_info': designer_info,
    }
    return render(request, 'process_csv.html', context)

def get_mounting_types_by_category(category_name):
    """Возвращает доступные типы монтажа для категории"""
    if 'петл' in category_name.lower():
        return Product.objects.filter(
            category__name__icontains='петл'
        ).exclude(mounting_type__isnull=True).exclude(mounting_type='').values_list(
            'mounting_type', flat=True
        ).distinct()
    elif 'направля' in category_name.lower():
        return Product.objects.filter(
            category__name__icontains='направля'
        ).exclude(mounting_type__isnull=True).exclude(mounting_type='').values_list(
            'mounting_type', flat=True
        ).distinct()
    return []

@login_required
def product_detail(request, pk):
    """Детальная информация о продукте"""
    product = get_object_or_404(Product.objects.select_related('category', 'brand', 'store'), id=pk)

    return render(request, 'product_detail.html', {
        'product': product
    })


def get_excel_data(request, pk):
    """Получить данные Excel для редактирования"""
    record = get_object_or_404(Record, id=pk)

    try:
        excel_file = RecordExcelFile.objects.get(record=record)
        file_path = excel_file.excel_file.path

        # Читаем Excel
        df = pd.read_excel(file_path, header=None)
        data = df.values.tolist()

        return JsonResponse({'data': data, 'success': True})

    except:
        return JsonResponse({'data': [], 'success': False})


def save_excel_data(request, pk):
    """Сохранить изменения из редактора"""
    record = get_object_or_404(Record, id=pk)

    if request.method == 'POST':
        try:
            excel_file = RecordExcelFile.objects.get(record=record)
            file_path = excel_file.excel_file.path

            # Получаем данные из запроса
            data = json.loads(request.POST.get('data', '[]'))

            # Сохраняем в Excel
            df = pd.DataFrame(data)
            df.to_excel(file_path, index=False, header=False)

            return JsonResponse({'success': True})

        except Exception as e:
            return JsonResponse({'success': False, 'error': str(e)})

    return JsonResponse({'success': False, 'error': 'Invalid method'})


def download_excel_file(request, pk):
    """Скачать Excel файл"""
    record = get_object_or_404(Record, id=pk)
    excel_file = get_object_or_404(RecordExcelFile, record=record)

    response = FileResponse(open(excel_file.excel_file.path, 'rb'))
    response['Content-Disposition'] = f'attachment; filename="комплектация_{pk}.xlsx"'
    return response


def calculate_record_total_components(record):
    """Вычисляет общую стоимость комплектующих для записи"""
    # Стоимость продуктов
    total_products = 0
    for record_product in record.recordproduct_set.all():
        if record_product.custom_price:
            total_products += float(record_product.custom_price) * record_product.quantity
        else:
            product_price = record_product.product.final_parsed_price
            total_products += float(product_price) * record_product.quantity
    
    # Стоимость плит
    plitas_total = 0
    for record_plita in record.recordplita_set.all():
        plitas_total += float(record_plita.plita.price_per_square_meter) * float(record_plita.quantity)
    
    return total_products + plitas_total

def calculate_record_total_expenses(record):
    """Вычисляет общую стоимость непланируемых расходов для записи"""
    return sum(float(expense.price) for expense in record.unplanned_expenses.all())

@login_required
def analytics_dashboard(request):
    """Панель аналитики с диаграммами и статистикой"""
    
    # Получаем все записи
    records = Record.objects.all()
    
    # 1. Анализ моржи по месяцам
    current_year = datetime.now().year
    margin_by_month = []
    for month in range(1, 13):
        month_records = records.filter(created_at__year=current_year, created_at__month=month)
        total_margin = 0
        for record in month_records:
            if record.contract_amount:
                total_components = calculate_record_total_components(record)
                margin = float(record.contract_amount) - total_components
                total_margin += margin
        margin_by_month.append({
            'month': calendar.month_name[month],
            'margin': round(total_margin, 2)
        })
    
    # 2. Распределение моржи между Юрой и Олегом
    yura_margin = 0
    oleg_margin = 0
    for record in records:
        if record.contract_amount:
            total_components = calculate_record_total_components(record)
            margin = float(record.contract_amount) - total_components
            if record.margin_yura and record.margin_oleg:
                # Если оба отмечены, делим пополам
                yura_margin += margin / 2
                oleg_margin += margin / 2
            elif record.margin_yura:
                yura_margin += margin
            elif record.margin_oleg:
                oleg_margin += margin
    
    margin_distribution = [
        {'name': 'Юра', 'value': round(yura_margin, 2)},
        {'name': 'Олег', 'value': round(oleg_margin, 2)}
    ]
    
    # 3. Статистика по статусам
    status_stats = []
    for status_code, status_name in Record.STATUS_CHOICES:
        count = records.filter(status=status_code).count()
        status_stats.append({
            'name': status_name,
            'count': count,
            'code': status_code
        })
    
    # 4. Топ проектировщиков по количеству заказов
    designer_stats = []
    designers = Designer.objects.all()
    for designer in designers:
        count = records.filter(designer=designer).count()
        if count > 0:
            designer_stats.append({
                'name': str(designer),
                'count': count
            })
    
    # 5. Средняя моржа по заказам
    total_margin = 0
    orders_with_margin = 0
    for record in records:
        if record.contract_amount:
            total_components = calculate_record_total_components(record)
            margin = float(record.contract_amount) - total_components
            total_margin += margin
            orders_with_margin += 1
    
    avg_margin = round(total_margin / orders_with_margin, 2) if orders_with_margin > 0 else 0
    
    # 6. Общая статистика
    total_orders = records.count()
    total_contract_amount = sum(float(record.contract_amount) for record in records if record.contract_amount)
    
    # Вычисляем общую стоимость комплектующих
    total_components_cost = 0
    for record in records:
        total_components_cost += calculate_record_total_components(record)
    
    # Вычисляем общие расходы
    total_expenses = 0
    for record in records:
        total_expenses += calculate_record_total_expenses(record)
    
    # 7. Анализ по дням недели (когда создаются заказы)
    weekday_stats = []
    weekdays = ['Понедельник', 'Вторник', 'Среда', 'Четверг', 'Пятница', 'Суббота', 'Воскресенье']
    for i, day in enumerate(weekdays):
        count = records.filter(created_at__week_day=i+2).count()  # Django week_day: 2=Monday, 7=Sunday
        weekday_stats.append({
            'day': day,
            'count': count
        })
    
    context = {
        'margin_by_month': margin_by_month,
        'margin_distribution': margin_distribution,
        'status_stats': status_stats,
        'designer_stats': designer_stats,
        'avg_margin': avg_margin,
        'total_orders': total_orders,
        'total_contract_amount': round(total_contract_amount, 2),
        'total_components_cost': round(total_components_cost, 2),
        'total_expenses': round(total_expenses, 2),
        'weekday_stats': weekday_stats,
        'current_year': current_year
    }
    
    return render(request, 'analytics/dashboard.html', context)